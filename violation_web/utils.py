"""
Utility functions used by the violation management web application.

This module centralises logic for importing the Excel workbook, computing
week numbers, formatting currency, building payment messages and generating
QR codes.  Keeping these routines separate from ``app.py`` improves
readability and makes it easier to test individual pieces of functionality.
"""

import base64
import io
import os
from datetime import datetime, date
from typing import Any, Dict, Iterable, Optional, Tuple, Type

import pandas as pd

try:
    import qrcode  # type: ignore
    _QRCODE_AVAILABLE = True
except ImportError:
    _QRCODE_AVAILABLE = False

try:
    from PIL import Image, ImageDraw, ImageFont  # type: ignore
    _PIL_AVAILABLE = True
except ImportError:
    _PIL_AVAILABLE = False


def compute_week_number(d: date) -> int:
    """
    Compute the ISO week number for a given date.

    The Excel workbook defines weeks similarly to ISO 8601 (weeks start on
    Monday and the first week of the year is the one containing the first
    Thursday).  Python's built‑in ``isocalendar`` method adheres to this
    specification, so we simply use it here.

    Args:
        d: A ``datetime.date`` instance.

    Returns:
        Integer week number (1–53).
    """
    return d.isocalendar().week


def format_currency(amount: int) -> str:
    """
    Format an integer amount (in VND) with thousand separators and suffix.
    For example, 10000 becomes ``"10,000 VND"``.
    """
    return f"{amount:,} VND".replace(',', '.')


def generate_payment_message(name: str, amount: int, codes: Iterable[str]) -> str:
    """
    Build the default transfer description for a payment.

    The description is required to follow the format
    ``Họ và tên + số tiền + mã lỗi``, where the amount uses dot
    separators for thousands and multiple codes are separated by
    commas and spaces.  For example:

    ``"Vũ Trần Đình Tâm 310.000 VP01, VP04, VP05"``

    Args:
        name: Full name of the student.
        amount: Amount still owed in VND.
        codes: Iterable of violation codes.

    Returns:
        A formatted transfer description string.
    """
    # format amount with dot separators and no currency suffix
    amount_str = f"{amount:,}".replace(',', '.')
    # join codes with comma and space
    code_str = ', '.join(codes)
    return f"{name} {amount_str} {code_str}"


def remove_violation_from_excel(record: Any, excel_path: str) -> None:
    """
    Remove a specific violation record from the original Excel workbook.

    This helper attempts to locate a row in the appropriate worksheet
    corresponding to the supplied ``ViolationRecord`` and delete it.  The
    match is based on the student name, date and error code.  Because
    the structure of the workbook is fixed, this function uses a
    simple heuristic: it iterates through data rows (identified by a
    numeric value in column A) and compares the name (column D), the
    date (column C) and the amount due (column H) to determine if a
    row should be removed.  If multiple rows match the criteria the
    first one encountered is deleted.  If no match is found the
    workbook is left unchanged.

    Args:
        record: A ``ViolationRecord`` instance from the database.
        excel_path: Path to the Excel workbook to update.
    """
    import datetime as _dt
    if not os.path.exists(excel_path):
        return
    try:
        import openpyxl  # type: ignore
    except ImportError:
        # If openpyxl is unavailable we cannot modify the file
        return
    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception:
        return
    sheet_name = record.sheet_name
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    # normalise the target name for comparison
    target_name = ' '.join([p.capitalize() for p in record.student_name.strip().split()])
    target_date_str = record.date.strftime('%Y-%m-%d') if isinstance(record.date, _dt.date) else str(record.date)
    # iterate over rows and identify the row to remove
    row_to_delete = None
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
        # Column A holds STT; skip non-numeric rows
        cell_val = row[0].value
        try:
            int(cell_val)
        except Exception:
            continue
        # Column C (index 2) holds date
        if len(row) > 2:
            date_cell = row[2].value
            # Normalise date cell to string for comparison
            date_val = None
            if isinstance(date_cell, _dt.datetime):
                date_val = date_cell.date().strftime('%Y-%m-%d')
            elif isinstance(date_cell, _dt.date):
                date_val = date_cell.strftime('%Y-%m-%d')
            elif isinstance(date_cell, (int, float)):
                # Excel serial date
                try:
                    excel_start = _dt.datetime(1899, 12, 30)
                    serial_date = excel_start + _dt.timedelta(days=int(date_cell))
                    date_val = serial_date.date().strftime('%Y-%m-%d')
                except Exception:
                    date_val = None
            elif isinstance(date_cell, str):
                date_val = date_cell.strip()
        else:
            date_val = None
        # Column D (index 3) holds name
        name_val = row[3].value if len(row) > 3 else None
        if not isinstance(name_val, str):
            continue
        current_name = ' '.join([p.capitalize() for p in name_val.strip().split()])
        # Column H (index 7) holds amount due; we compare to ensure correct row
        amount_val = row[7].value if len(row) > 7 else None
        try:
            current_amount = int(str(amount_val).replace(',', '').replace('.', '')) if amount_val not in (None, '') else 0
        except Exception:
            current_amount = 0
        # Compare name, date and amount
        if current_name == target_name and date_val == target_date_str and current_amount == record.amount_due:
            row_to_delete = idx
            break
    if row_to_delete:
        try:
            ws.delete_rows(row_to_delete, 1)
            wb.save(excel_path)
        except Exception:
            pass


def generate_qr_code_base64(data: str) -> str:
    """
    Generate a QR code representing the given data and return it as a
    base64‑encoded PNG.  If the ``qrcode`` library is not available the
    function falls back to producing a simple placeholder image so that
    templates can still display something.

    Args:
        data: The text to encode in the QR code.

    Returns:
        A base64 data URI ready to be embedded in an HTML ``<img>`` tag.
    """
    if _QRCODE_AVAILABLE:
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
    elif _PIL_AVAILABLE:
        # Fallback: create a blank image with the data printed in the centre
        img = Image.new('RGB', (200, 200), color=(255, 255, 255))
        draw = ImageDraw.Draw(img)
        try:
            # Use a default font
            font = ImageFont.load_default()
        except Exception:
            font = None
        text = 'QR'  # Indicate that this is a placeholder
        w, h = draw.textsize(text, font=font)
        draw.text(((200 - w) / 2, (200 - h) / 2), text, fill=(0, 0, 0), font=font)
    else:
        # If neither qrcode nor PIL is available return an empty string
        return ''
    buffered = io.BytesIO()
    img.save(buffered, format="PNG")
    encoded = base64.b64encode(buffered.getvalue()).decode('ascii')
    return f"data:image/png;base64,{encoded}"


def compute_custom_week(d: date) -> int:
    """
    Compute a custom week number starting from 8 September 2025 (week 1)
    and skipping the two‑week Tết break from 15 February 2026 to
    28 February 2026 inclusive.  Weeks start on Monday and each week
    covers a 7‑day period.  If the date falls within the break it is
    assigned to the week immediately preceding the break.

    Args:
        d: The date to calculate the week for.

    Returns:
        The custom week number (starting at 1).
    """
    # Define start and break periods
    start_date = date(2025, 9, 8)  # Monday 8/9/2025 is week 1
    break_start = date(2026, 2, 15)
    break_end = date(2026, 2, 28)
    if d < start_date:
        return 1
    # If within the break, treat as previous day just before break
    if break_start <= d <= break_end:
        d_effective = break_start - pd.Timedelta(days=1)
    else:
        d_effective = d
    # Compute days since start, adjusting for break period
    delta_days = (d_effective - start_date).days
    # If the date is after the break, subtract the length of the break
    if d_effective > break_end:
        delta_days -= (break_end - break_start).days + 1
    week_num = delta_days // 7 + 1
    return week_num


def import_excel_if_needed(
    excel_path: str,
    db,
    ErrorCode: Type,
    ViolationRecord: Type,
) -> None:
    """
    Read the provided Excel workbook and populate the ``violation_records`` table.

    This function is only called if there are no existing ``ViolationRecord``
    entries.  It imports data from the known worksheets and maps them to
    database rows.  Columns from the workbook are normalised and missing
    values are replaced with defaults defined in the ``ErrorCode`` table.

    Args:
        excel_path: Absolute path to the ``.xlsm`` file.
        db: SQLAlchemy database instance.
        ErrorCode: SQLAlchemy model for error codes.
        ViolationRecord: SQLAlchemy model for violation records.
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    # define mapping from sheet names to error codes
    sheet_map: Dict[str, Tuple[str, int]] = {
        'NHAT_KI_DI_MUON': ('VP01', 10000),
        'NG_LA': ('VP02', 0),
        'DOI_CHO': ('VP03', 10000),
        'QUEN_DDHT': ('VP04', 10000),
        'NGU_TRONG_GIO': ('VP05', 10000),
        'NGHI_HOC': ('VP06', 30000),
    }
    # iterate through each sheet of interest
    xls = pd.ExcelFile(excel_path)
    for sheet_name, (code, default_amount) in sheet_map.items():
        if sheet_name not in xls.sheet_names:
            continue
        # load sheet; leave header as None so we can process manually
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
        # iterate over rows and extract records
        for idx, row in df.iterrows():
            # skip header rows or separators
            stt = row.iloc[0]
            if pd.isna(stt):
                continue
            # ignore non‑numeric identifiers
            try:
                int(stt)
            except Exception:
                continue
            # parse week
            week_val = row.iloc[1]
            try:
                week = int(week_val) if not pd.isna(week_val) else None
            except Exception:
                week = None
            # parse date
            date_val = row.iloc[2]
            record_date: Optional[date] = None
            if isinstance(date_val, (datetime, pd.Timestamp)):
                record_date = date_val.date()
            elif isinstance(date_val, str):
                try:
                    record_date = datetime.strptime(date_val, '%Y-%m-%d').date()
                except Exception:
                    pass
            # parse name
            name = row.iloc[3] if len(row) > 3 else None
            if pd.isna(name):
                continue
            name = str(name).strip()
            # parse reason (column 4)
            reason = None
            if len(row) > 4:
                val = row.iloc[4]
                if isinstance(val, str):
                    reason = val.strip()
            # amount paid (column 5)
            amount_paid = 0
            if len(row) > 5:
                val = row.iloc[5]
                if isinstance(val, (int, float)) and not pd.isna(val):
                    amount_paid = int(val)
            # payment date (column 6)
            payment_date = None
            if len(row) > 6:
                val = row.iloc[6]
                if isinstance(val, (datetime, pd.Timestamp)):
                    payment_date = val.date()
                elif isinstance(val, (float, int)):
                    # Excel serial date number
                    try:
                        payment_date = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(val) - 2).date()
                    except Exception:
                        payment_date = None
            # amount due (column 7)
            amount_due = default_amount
            if len(row) > 7:
                val = row.iloc[7]
                if isinstance(val, (int, float)) and not pd.isna(val):
                    amount_due = int(val)
            # notes (column 8)
            notes = None
            if len(row) > 8:
                val = row.iloc[8]
                if isinstance(val, str):
                    notes = val.strip()
            # fallback week if missing: compute from date
            if week is None and record_date is not None:
                week = compute_week_number(record_date)
            # ensure we have date
            if record_date is None:
                continue
            # normalise names (capitalisation) to match DS_LOP
            name = ' '.join([part.capitalize() for part in name.split()])
            rec = ViolationRecord(
                sheet_name=sheet_name,
                week=week or 0,
                date=record_date,
                student_name=name,
                error_code=code,
                reason=reason,
                amount_due=amount_due,
                amount_paid=amount_paid,
                payment_date=payment_date,
                notes=notes,
            )
            db.session.add(rec)
    # commit occurs outside this function


def get_ds_lop_names(excel_path: str) -> list:
    """
    Extract a list of student names from the ``DS_LOP`` worksheet of
    the given workbook.  Names are normalised to title case and
    duplicates are removed.

    Args:
        excel_path: Absolute path to the Excel workbook.

    Returns:
        A sorted list of unique names.  If the worksheet does not exist
        or the file cannot be read an empty list is returned.
    """
    if not os.path.exists(excel_path):
        return []
    try:
        df = pd.read_excel(excel_path, sheet_name='DS_LOP', header=None)
    except Exception:
        return []
    names = []
    for idx, row in df.iterrows():
        # second column (index 1) contains names after STT
        stt_val = row.iloc[0]
        # skip header or empty rows
        if isinstance(stt_val, str) and stt_val.strip().lower() == 'stt':
            continue
        name_val = row.iloc[1] if len(row) > 1 else None
        if isinstance(name_val, str) and name_val.strip():
            name = ' '.join([part.capitalize() for part in name_val.strip().split()])
            names.append(name)
    # remove duplicates and sort
    unique_names = sorted(set(names))
    return unique_names


def update_excel_payment(student_name: str, excel_path: str) -> None:
    """
    Mark all outstanding debts for the specified student as paid in the
    original Excel workbook.  For each relevant worksheet the function
    will move the value from ``Số Tiền Chưa Nộp`` into ``Nộp Tiền`` and
    set the unpaid amount to zero.  It also records the current date in
    the ``Ngày Nộp`` column.

    This function makes a best‑effort attempt based on the fixed
    structure of the workbook.  It does not handle merged cells or
    complicated macros; it simply writes values into the existing cells.

    Args:
        student_name: Normalised name (title case) of the student.
        excel_path: Path to the original Excel file to update.
    """
    if not os.path.exists(excel_path):
        return
    try:
        import openpyxl  # defer import to avoid dependency when unused
    except ImportError:
        return
    # sheets mapping to update
    sheet_names = [
        'NHAT_KI_DI_MUON', 'NG_LA', 'DOI_CHO', 'QUEN_DDHT', 'NGU_TRONG_GIO', 'NGHI_HOC'
    ]
    wb = openpyxl.load_workbook(excel_path)
    today_str = datetime.today().strftime('%Y-%m-%d')
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # iterate rows; header is in first few rows; data rows contain numeric STT in column A
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            stt_cell = row[0]
            try:
                stt = int(stt_cell.value)
            except Exception:
                continue
            # get name (column index 3) if present
            if len(row) < 4:
                continue
            name_cell = row[3]
            name_val = name_cell.value
            if not isinstance(name_val, str):
                continue
            # normalise both names for comparison
            target_name = ' '.join([part.capitalize() for part in name_val.strip().split()])
            if target_name != student_name:
                continue
            # get unpaid amount from column index 7 (H column) if exists
            if len(row) < 8:
                continue
            unpaid_cell = row[7]
            amount_unpaid = 0
            try:
                if unpaid_cell.value is not None:
                    amount_unpaid = int(str(unpaid_cell.value).replace(',', '').replace('.', ''))
            except Exception:
                amount_unpaid = 0
            if amount_unpaid <= 0:
                continue
            # update paid amount (column 5, index 5)
            if len(row) > 5:
                paid_cell = row[5]
                prev_paid = 0
                try:
                    if paid_cell.value is not None:
                        prev_paid = int(str(paid_cell.value).replace(',', '').replace('.', ''))
                except Exception:
                    prev_paid = 0
                new_paid = prev_paid + amount_unpaid
                paid_cell.value = f"{new_paid}"
            # update payment date (column 6, index 6)
            if len(row) > 6:
                date_cell = row[6]
                date_cell.value = today_str
            # reset unpaid amount
            unpaid_cell.value = 0
    wb.save(excel_path)


def append_violation_to_excel(record: Any, excel_path: str) -> None:
    """
    Append a new violation record to the appropriate sheet in the original
    Excel workbook.  This is used when users add new violations via
    the web interface.  The function attempts to preserve the same
    column ordering as the imported data: STT, Tuần, Ngày, Họ tên,
    Lý do, Số tiền đã nộp, Ngày nộp, Số tiền phải nộp, Ghi chú.

    Args:
        record: A ViolationRecord instance (or object with similar
            attributes) containing the data to append.
        excel_path: Path to the original Excel file to update.
    """
    if not os.path.exists(excel_path):
        return
    try:
        import openpyxl
    except ImportError:
        return
    wb = openpyxl.load_workbook(excel_path)
    sheet_name = getattr(record, 'sheet_name', None)
    if not sheet_name or sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    # Determine the next STT by scanning the first column for the last numeric value
    last_stt = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cell = row[0]
        try:
            val = int(cell.value)
            if val > last_stt:
                last_stt = val
        except Exception:
            continue
    new_stt = last_stt + 1
    # Build row values
    week = getattr(record, 'week', None) or ''
    date_val = getattr(record, 'date', None)
    date_str = ''
    if isinstance(date_val, (datetime, date)):
        date_str = date_val.strftime('%Y-%m-%d')
    student = getattr(record, 'student_name', '')
    reason = getattr(record, 'reason', None) or ''
    amount_paid = getattr(record, 'amount_paid', 0) or 0
    payment_date = getattr(record, 'payment_date', None)
    payment_date_str = ''
    if isinstance(payment_date, (datetime, date)):
        payment_date_str = payment_date.strftime('%Y-%m-%d')
    amount_due = getattr(record, 'amount_due', 0) or 0
    notes = getattr(record, 'notes', None) or ''
    new_row = [new_stt, week, date_str, student, reason, amount_paid, payment_date_str, amount_due, notes]
    ws.append(new_row)
    wb.save(excel_path)